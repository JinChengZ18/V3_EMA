"""Cross-version diff for regions reports.

Reads two regions xlsx reports and computes per-state added/removed/changed.
Captures dynamic resource columns (whose headers are localized building names,
not in the i18n table) so they can participate in the diff and the writer can
reproduce the same column layout as the main report."""
from __future__ import annotations
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

from ..i18n import label_to_key
from ..util.logging import get_logger
from .diff import ChangedRow, _is_changed, _read_info_sheet  # reuse helpers

log = get_logger()

REGION_KEY_FIELDS = ("state_id",)
REGION_NUMERIC_FIELDS = (
    "arable_land", "capped_total",
    "total_capacity", "provinces", "numeric_id",
)
# Canonical ID-like text fields — language-independent, safe to compare.
REGION_CANONICAL_FIELDS = ("traits_ids", "strat_id")
# Localized display fields — included in changed-row context but NOT compared
# (so cross-language diffs don't false-positive on every label).
REGION_TEXT_FIELDS = (
    "state", "strategic_region", "arable_buildings",
    "capped_resources", "discoverable", "known_resources",
    "traits", "trait_modifiers", "subsistence",
)
# Field-name prefix used internally to mark dynamic resource columns whose
# headers are raw localized building names (e.g. "煤矿"). Stored as `res_煤矿`.
RES_PREFIX = "res_"

_REGION_SHEET_KEYS = frozenset({
    "rsheet_overview", "sheet_overview",
    "rbucket_western_europe", "rbucket_southern_europe",
    "rbucket_northern_europe", "rbucket_eastern_europe",
    "rbucket_north_america", "rbucket_central_america", "rbucket_south_america",
    "rbucket_africa", "rbucket_middle_east", "rbucket_central_asia",
    "rbucket_india", "rbucket_east_asia", "rbucket_southeast_asia",
    "rbucket_oceania", "rbucket_other",
    "bucket_other",
})
_OVERVIEW_KEYS = frozenset({"rsheet_overview", "sheet_overview"})


@dataclass
class RegionsSnapshot:
    meta: dict[str, str] = field(default_factory=dict)
    states: dict[tuple, dict[str, object]] = field(default_factory=dict)
    # Column layout from the primary (overview) sheet — preserves header order
    # so the diff writer can reproduce the same column arrangement as the main
    # report. Each entry: (canonical_field_name, header_text).
    column_layout: list[tuple[str, str]] = field(default_factory=list)


@dataclass
class RegionsDiffResult:
    old_meta: dict[str, str]
    new_meta: dict[str, str]
    added: list[dict[str, object]] = field(default_factory=list)
    removed: list[dict[str, object]] = field(default_factory=list)
    changed: list[ChangedRow] = field(default_factory=list)
    # Column layout to use when rendering — taken from the new snapshot.
    column_layout: list[tuple[str, str]] = field(default_factory=list)


def _header_to_field(header) -> str | None:
    """Map an xlsx header text to a canonical field name.

    Strips the i18n key prefix (rcol_/col_/ccol_) when found. For unmapped
    headers (typically dynamic resource columns localized as building names),
    returns `res_<header>` so they get diffed too. Empty / None headers
    return None and are skipped.
    """
    if header is None or str(header).strip() == "":
        return None
    ui_k = label_to_key(header)
    if ui_k:
        if ui_k.startswith("rcol_"):
            return ui_k[5:]
        if ui_k.startswith("col_"):
            return ui_k[4:]
        if ui_k.startswith("ccol_"):
            return ui_k[5:]
    return f"{RES_PREFIX}{header}"


def _read_region_sheet(
    wb, sheet_name: str,
) -> tuple[dict[tuple, dict[str, object]], list[tuple[str, str]]]:
    """Returns (data, layout). Data is {state_key: {field: value}}; layout is
    [(field_name, header_text), ...] preserving sheet column order."""
    if sheet_name not in wb.sheetnames:
        return {}, []
    ws = wb[sheet_name]
    headers = [c.value for c in ws[1]]

    layout: list[tuple[str, str]] = []
    canonical: list[str | None] = []
    for h in headers:
        f = _header_to_field(h)
        canonical.append(f)
        if f is not None:
            layout.append((f, str(h)))

    col_for: dict[str, int] = {}
    for idx, f in enumerate(canonical):
        if f and f not in col_for:
            col_for[f] = idx

    if "state_id" not in col_for:
        log.warning("Regions sheet %s has no state_id column (headers=%s)",
                    sheet_name, headers)
        return {}, []

    state_idx = col_for["state_id"]

    out: dict[tuple, dict[str, object]] = {}
    skip_first_data_row = True   # row 2 is the totals row in regions reports
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        # Skip totals row (state_id will be None or a localization label like "合计")
        if skip_first_data_row:
            skip_first_data_row = False
            sid_val = row[state_idx]
            if sid_val is None or not str(sid_val).startswith("STATE_"):
                continue
        state_id = row[state_idx]
        if state_id is None or not str(state_id).startswith("STATE_"):
            continue
        key = (str(state_id),)
        # Capture every value whose column had a canonical mapping.
        out[key] = {col_for_key: row[col_idx]
                    for col_for_key, col_idx in col_for.items()}
    return out, layout


def read_regions_report(path: Path) -> RegionsSnapshot:
    wb = load_workbook(path, read_only=True, data_only=True)
    snap = RegionsSnapshot(meta=_read_info_sheet(wb))

    # Read overview first (records the canonical layout); bucket sheets override.
    sheet_priority = []
    for name in wb.sheetnames:
        k = label_to_key(name)
        if k in _REGION_SHEET_KEYS:
            sheet_priority.append((0 if k in _OVERVIEW_KEYS else 1, name))
    sheet_priority.sort()

    states: dict[tuple, dict[str, object]] = {}
    layout: list[tuple[str, str]] = []
    for prio, name in sheet_priority:
        sheet_data, sheet_layout = _read_region_sheet(wb, name)
        states.update(sheet_data)
        if not layout and prio == 0:
            layout = sheet_layout
    snap.states = states
    snap.column_layout = layout
    wb.close()
    return snap


def diff_regions_snapshots(
    old: RegionsSnapshot,
    new: RegionsSnapshot,
    *,
    eps_abs: float = 0.01,
    eps_rel: float = 0.005,
) -> RegionsDiffResult:
    res = RegionsDiffResult(
        old_meta=old.meta, new_meta=new.meta,
        column_layout=new.column_layout or old.column_layout,
    )
    old_keys = set(old.states)
    new_keys = set(new.states)
    for k in sorted(new_keys - old_keys):
        res.added.append({"_key": k, **new.states[k]})
    for k in sorted(old_keys - new_keys):
        res.removed.append({"_key": k, **old.states[k]})

    # Comparable fields: numeric + canonical ID fields + dynamic res_* columns.
    # We collect the union of res_* keys actually present in both reports.
    res_fields: set[str] = set()
    for d in (*old.states.values(), *new.states.values()):
        res_fields.update(f for f in d if f.startswith(RES_PREFIX))

    numeric_compare = set(REGION_NUMERIC_FIELDS) | res_fields
    id_compare = set(REGION_CANONICAL_FIELDS)

    for k in sorted(old_keys & new_keys):
        o, n = old.states[k], new.states[k]
        deltas: dict[str, tuple] = {}
        for f in numeric_compare:
            ov, nv = o.get(f), n.get(f)
            if _is_changed(ov, nv, eps_abs, eps_rel):
                deltas[f] = (ov, nv)
        for f in id_compare:
            ov, nv = o.get(f), n.get(f)
            if (ov or "") != (nv or ""):
                deltas[f] = (ov, nv)
        if deltas:
            # text_fields carries the new row's identifying text for context.
            text = {f: str(n.get(f, "") or "") for f in REGION_TEXT_FIELDS}
            res.changed.append(ChangedRow(key=k, text_fields=text, deltas=deltas))
    return res
