from __future__ import annotations
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..analysis.regions import (
    REGION_BUCKET_ORDER,
    REGION_COLUMN_KEYS,
    RegionRow,
    discover_dynamic_resources,
    get_row_value,
    make_region_columns,
    make_region_columns_with_dynamic,
)
from ..i18n import UI, get_ui
from ..model import GameData
from .xlsx_writer import (
    BAND_FILL,
    FONT_FAMILY,
    HEADER_BORDER,
    LEFT,
    ReportMeta,
    TEXT_MUTED,
    TEXT_PRIMARY,
    _cell_value,
    _style_data_sheet,
)

TOTALS_FILL = PatternFill("solid", fgColor="E0E7FF")  # indigo-100, distinct from header

# Static numeric keys eligible for the totals row (sum-able across states).
_NUMERIC_TOTAL_KEYS = {
    "arable_land", "capped_total", "total_capacity", "provinces",
}

# Column-key categorization specific to the regions report.
_REGION_NUM_INT_KEYS = {
    "arable_land", "capped_total", "resource_kinds",
    "total_capacity", "provinces", "numeric_id",
}
_REGION_TEXT_KEYS = {
    "state", "strategic_region", "arable_buildings", "capped_resources",
    "discoverable", "known_resources", "traits", "trait_modifiers",
    "subsistence",
}
_REGION_ID_KEYS = {"state_id", "traits_ids", "strat_id"}

_REGION_HEADER_ONLY_WIDTH_KEYS = {"state_id", "traits_ids", "strat_id"}

_REGION_COL_WIDTH_CAP = {
    "state": 18,
    "strategic_region": 18,
    "arable_land": 12,
    "arable_buildings": 24,
    "capped_total": 12,
    "capped_resources": 38,
    "discoverable": 28,
    "known_resources": 24,
    "resource_kinds": 12,
    "total_capacity": 14,
    "traits": 28,
    "trait_modifiers": 60,
    "subsistence": 18,
    "provinces": 10,
    "numeric_id": 10,
}

_REGION_MIN_WIDTH = {
    "arable_land": 10,
    "capped_total": 12,
    "total_capacity": 14,
    "resource_kinds": 12,
    "provinces": 10,
    "numeric_id": 10,
}

_REGION_CONDITIONAL_KEYS = ("arable_land", "capped_total", "total_capacity")


def _patch_writer_classifiers():
    """Inject our region-specific column classifications into the shared
    `_style_data_sheet` helper. The helper reads several module-level sets
    (NUM_INT_KEYS / NUM_VAL_KEYS / NUM_RATE_KEYS / TEXT_KEYS / ID_KEYS /
    HEADER_ONLY_WIDTH_KEYS / MIN_WIDTH / COL_WIDTH_CAP); we extend them in
    place so region columns get the right styling. Idempotent."""
    from . import xlsx_writer as xw
    xw.NUM_INT_KEYS.update(_REGION_NUM_INT_KEYS)
    xw.TEXT_KEYS.update(_REGION_TEXT_KEYS)
    xw.ID_KEYS.update(_REGION_ID_KEYS)
    xw.HEADER_ONLY_WIDTH_KEYS.update(_REGION_HEADER_ONLY_WIDTH_KEYS)
    xw.COL_WIDTH_CAP.update(_REGION_COL_WIDTH_CAP)
    xw.MIN_WIDTH.update(_REGION_MIN_WIDTH)


def _write_info_sheet(wb: Workbook, meta: ReportMeta, ui: UI) -> None:
    ws = wb.create_sheet(ui["sheet_info"], 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 38

    ws["A1"] = ui["meta_title"]
    ws["A1"].font = Font(name=FONT_FAMILY, size=14, bold=True, color=TEXT_PRIMARY)
    ws.merge_cells("A1:B1")

    rows: list[tuple[str, str]] = [
        (ui["meta_game_version"], meta.game_version or ui["meta_unknown"]),
        (ui["meta_raw_version"],  meta.raw_version or ui["meta_unknown"]),
        (ui["meta_generated_at"], meta.generated_at),
        (ui["meta_tool_version"], f"V3_EAT {meta.tool_version}"),
        (ui["meta_data_lang"],    meta.data_lang or ui["meta_unknown"]),
        (ui["meta_ui_lang"],      meta.ui_lang or ui["meta_unknown"]),
    ]
    count_keys = (
        ("rcount_states",            "states"),
        ("rcount_state_traits",      "state_traits"),
        ("rcount_strategic_regions", "strategic_regions"),
    )
    for ui_key, count_key in count_keys:
        if count_key in meta.counts:
            rows.append((ui[ui_key], f"{meta.counts[count_key]:,}"))

    for i, (label, value) in enumerate(rows, start=3):
        kc = ws.cell(row=i, column=1, value=label)
        kc.font = Font(name=FONT_FAMILY, size=10, color=TEXT_MUTED)
        kc.alignment = LEFT
        vc = ws.cell(row=i, column=2, value=value)
        vc.font = Font(name=FONT_FAMILY, size=10, color=TEXT_PRIMARY, bold=True)
        vc.alignment = LEFT
        if i % 2 == 1:
            kc.fill = BAND_FILL
            vc.fill = BAND_FILL


def _sum_numeric(rows: list[RegionRow], key: str):
    total = 0.0
    seen = False
    for r in rows:
        v = get_row_value(r, key)
        if isinstance(v, (int, float)):
            total += v
            seen = True
    if not seen:
        return None
    return int(total) if total == int(total) else round(total, 2)


def _write_regions_sheet(
    wb: Workbook, sheet_name: str, rows: list[RegionRow], ui: UI,
    *, resource_ids: list[str], loc_get,
) -> None:
    ws = wb.create_sheet(sheet_name)
    columns = make_region_columns_with_dynamic(ui, resource_ids, loc_get)
    headers = [label for _, label in columns]
    keys = [k for k, _ in columns]
    ws.append(headers)

    # Totals row (immediately under the header). Sums numeric + dynamic
    # cap_/pot_ columns; leaves text cells empty; tags the state column "合计".
    totals: list = []
    for k, _label in columns:
        if k == "state":
            totals.append(ui["totals_label"])
        elif k in _NUMERIC_TOTAL_KEYS or k.startswith("res_"):
            totals.append(_sum_numeric(rows, k))
        else:
            totals.append(None)
    ws.append(totals)

    for r in rows:
        ws.append([_cell_value(get_row_value(r, k)) for k in keys])

    _style_data_sheet(
        ws, rows, headers, keys,
        building_col_key="state",                # state name = anchor column
        conditional_keys=_REGION_CONDITIONAL_KEYS,
        fill_anchor_col=False,                   # but no fill — keeps the color domain neutral
        data_row_offset=1,                       # row 2 is the totals row (caller-prepended)
    )

    # Style the totals row (row 2): bold, indigo-100 fill, top+bottom border.
    totals_row_idx = 2
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=totals_row_idx, column=col_idx)
        cell.fill = TOTALS_FILL
        cell.font = Font(name=FONT_FAMILY, size=10, bold=True, color=TEXT_PRIMARY)
        cell.border = HEADER_BORDER
        if isinstance(cell.value, (int, float)):
            cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)
            cell.number_format = "#,##0"
        else:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Freeze pane so both header AND totals row stay visible on scroll.
    if "state" in keys:
        anchor_col = get_column_letter(keys.index("state") + 2)
        ws.freeze_panes = f"{anchor_col}3"


def _write_maps_sheet(wb: Workbook, ui: UI, images: list[tuple[str, Path]]) -> None:
    """Embed pre-rendered choropleth PNGs into a single 'Resource Maps' sheet,
    stacked vertically with a bold label above each."""
    from openpyxl.drawing.image import Image as XLImage

    ws = wb.create_sheet(ui["map_sheet"])
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 18
    r = 1
    for label, img_path in images:
        if not img_path.exists():
            continue
        c = ws.cell(row=r, column=1, value=label)
        c.font = Font(name=FONT_FAMILY, size=12, bold=True, color=TEXT_PRIMARY)
        c.alignment = LEFT
        try:
            xim = XLImage(str(img_path))
        except Exception:  # pragma: no cover - missing PIL/decoder
            continue
        ws.add_image(xim, f"A{r + 1}")
        # ~18 px per default row; leave the image's height plus a gap.
        r += int(getattr(xim, "height", 600) / 18) + 4


def write_regions_xlsx(
    rows: Iterable[RegionRow],
    path: Path,
    *,
    meta: ReportMeta | None = None,
    ui: UI | None = None,
    game: GameData | None = None,
    map_images: list[tuple[str, Path]] | None = None,
) -> int:
    _patch_writer_classifiers()
    rows_list = list(rows)
    path.parent.mkdir(parents=True, exist_ok=True)

    if ui is None:
        ui = get_ui(meta.data_lang if meta else "simp_chinese")
    if meta is None:
        meta = ReportMeta(generated_at=datetime.now().isoformat(timespec="seconds"))

    # Discover dynamic column set from the rows themselves (so callers don't
    # need to pass `game` if they already built rows). Falls back to game.
    if game is not None:
        resource_ids = discover_dynamic_resources(game)
    else:
        seen: set[str] = set()
        for r in rows_list:
            seen.update(r.res_by_id.keys())
        resource_ids = sorted(seen)

    # Localizer for the dynamic column headers (building names).
    loc = getattr(game, "loc", None) if game is not None else None

    def loc_get(k: str) -> str:
        return loc.get_clean(k) if loc is not None else k

    # Allow xlsx_writer's _style_data_sheet to recognize res_ keys.
    from . import xlsx_writer as xw
    for bld in resource_ids:
        xw.NUM_INT_KEYS.add(f"res_{bld}")
        xw.COL_WIDTH_CAP[f"res_{bld}"] = 10
        xw.MIN_WIDTH[f"res_{bld}"] = 8

    wb = Workbook()
    default = wb.active
    wb.remove(default)

    _write_info_sheet(wb, meta, ui)

    by_bucket: dict[str, list[RegionRow]] = defaultdict(list)
    for r in rows_list:
        by_bucket[r.bucket or "other"].append(r)

    sheet_kwargs = dict(resource_ids=resource_ids, loc_get=loc_get)

    if rows_list:
        # Overview sorted by capped_total too (for consistency with bucket sheets).
        rows_for_overview = sorted(
            rows_list, key=lambda r: r.capped_total, reverse=True,
        )
        _write_regions_sheet(wb, ui["rsheet_overview"], rows_for_overview, ui, **sheet_kwargs)

    for bucket in REGION_BUCKET_ORDER:
        bucket_rows = by_bucket.get(bucket, [])
        if not bucket_rows:
            continue
        # Sort by capped resources (more relevant than total_capacity, since
        # arable land doesn't matter much in early game).
        bucket_rows.sort(key=lambda r: r.capped_total, reverse=True)
        _write_regions_sheet(wb, ui[f"rbucket_{bucket}"], bucket_rows, ui, **sheet_kwargs)

    if map_images:
        _write_maps_sheet(wb, ui, map_images)

    wb.properties.title = f"V3_EAT Regions Report — V3 {meta.raw_version or '?'}"
    wb.properties.creator = "V3_EAT"
    wb.properties.description = (
        f"Game: {meta.game_version} | Tool: V3_EAT {meta.tool_version} | "
        f"Generated: {meta.generated_at} | Data: {meta.data_lang} | UI: {meta.ui_lang}"
    )

    wb.save(path)
    return len(rows_list)
