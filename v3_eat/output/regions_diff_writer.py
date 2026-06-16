"""Diff workbook writer for regions reports.

Layout principle (per user feedback): the diff sheets reuse the SAME column
order as the main regions report (state → strategic_region → arable_land →
total_capacity → capped_total → per-resource columns → traits/text → IDs).

For changed rows: each cell shows the **delta in place** — no side-by-side
old/new columns. For numeric columns, the delta is `new - old`, color-coded
(green/red). For text columns, shows new value when changed. Unchanged
cells stay blank, so the user immediately sees which fields shifted."""
from __future__ import annotations
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..analysis.regions_diff import (
    REGION_KEY_FIELDS,
    RegionsDiffResult,
)
from ..i18n import UI, get_ui
from .diff_writer import (
    ADDED_FILL,
    BAND_FILL,
    BODY_FONT,
    DELTA_NEGATIVE,
    DELTA_POSITIVE,
    FONT_FAMILY,
    HEADER_FILL,
    HEADER_FONT,
    LEFT,
    MUTED_FONT,
    REMOVED_FILL,
    RIGHT,
    TEXT_PRIMARY,
    TITLE_FONT,
    _style_header,
)

# Helpful per-key column-width hints for the diff layout.
_WIDTH_HINTS = {
    "state": 16, "strategic_region": 16,
    "arable_land": 10, "arable_buildings": 22,
    "total_capacity": 12, "capped_total": 12,
    "traits": 24, "trait_modifiers": 50, "subsistence": 16,
    "provinces": 10, "numeric_id": 10,
    "capped_resources": 32, "discoverable": 24, "known_resources": 22,
    "state_id": 26, "traits_ids": 26, "strat_id": 22,
}


def _meta(diff: RegionsDiffResult, ui: UI, key_keys: tuple[str, ...]) -> str:
    for k in key_keys:
        v = diff.old_meta.get(k) or diff.new_meta.get(k)
        if v:
            return v
    return ui["meta_unknown"]


def _write_info_sheet(wb: Workbook, diff: RegionsDiffResult, ui: UI) -> None:
    ws = wb.create_sheet(ui["sheet_info"], 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 38

    ws["A1"] = ui["diff_title"]
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:B1")

    rows = [
        (ui["diff_old_game"],
         diff.old_meta.get(ui["meta_game_version"]) or diff.old_meta.get("游戏版本")
            or diff.old_meta.get("Game Version") or ui["meta_unknown"]),
        (ui["diff_old_generated"],
         diff.old_meta.get(ui["meta_generated_at"]) or diff.old_meta.get("生成时间")
            or diff.old_meta.get("Generated At") or ui["meta_unknown"]),
        (ui["diff_new_game"],
         diff.new_meta.get(ui["meta_game_version"]) or diff.new_meta.get("游戏版本")
            or diff.new_meta.get("Game Version") or ui["meta_unknown"]),
        (ui["diff_new_generated"],
         diff.new_meta.get(ui["meta_generated_at"]) or diff.new_meta.get("生成时间")
            or diff.new_meta.get("Generated At") or ui["meta_unknown"]),
        (ui["diff_compared_at"], datetime.now().isoformat(timespec="seconds")),
        ("", ""),
        ("Added / 新增",   str(len(diff.added))),
        ("Removed / 移除", str(len(diff.removed))),
        ("Changed / 变更", str(len(diff.changed))),
    ]
    for i, (k, v) in enumerate(rows, start=3):
        kc = ws.cell(row=i, column=1, value=k)
        vc = ws.cell(row=i, column=2, value=v)
        kc.font = MUTED_FONT
        vc.font = Font(name=FONT_FAMILY, size=10, bold=True, color=TEXT_PRIMARY)
        kc.alignment = LEFT
        vc.alignment = LEFT
        if i % 2 == 1:
            kc.fill = BAND_FILL
            vc.fill = BAND_FILL


def _resolve_layout(diff: RegionsDiffResult) -> list[tuple[str, str]]:
    """Use the snapshot's column layout. If empty (e.g. legacy report missing
    the dynamic columns), fall back to a minimal static layout."""
    if diff.column_layout:
        return diff.column_layout
    # Minimal fallback: just key + a few columns
    return [
        ("state_id", "State ID"),
        ("state", "State"),
        ("arable_land", "Arable"),
        ("capped_total", "Cap Total"),
        ("total_capacity", "Total Cap"),
    ]


def _is_numeric_field(field_name: str) -> bool:
    """Numeric column: numeric base fields OR dynamic res_* columns."""
    if field_name.startswith("res_"):
        return True
    return field_name in {
        "arable_land", "capped_total", "total_capacity",
        "provinces", "numeric_id",
    }


def _style_data_header(ws, n_cols: int) -> None:
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24


def _apply_widths(ws, layout: list[tuple[str, str]]) -> None:
    for idx, (field_name, header) in enumerate(layout, start=1):
        letter = get_column_letter(idx)
        hint = _WIDTH_HINTS.get(field_name)
        if hint:
            ws.column_dimensions[letter].width = hint
        else:
            ws.column_dimensions[letter].width = max(10, min(20, len(str(header)) * 2 + 2))


def _write_full_rows_sheet(
    wb: Workbook,
    sheet_name: str,
    rows: list[dict],
    layout: list[tuple[str, str]],
    *,
    fill_color: str,
) -> None:
    """For added/removed sheets: write the full row data using the main-table
    column order. The state-id column gets a colored fill for visual signal."""
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    headers = [h for _, h in layout]
    ws.append(headers)
    _style_data_header(ws, len(headers))

    fill = PatternFill("solid", fgColor=fill_color)
    for ridx, row in enumerate(rows, start=2):
        values = [row.get(field_name) for field_name, _ in layout]
        ws.append(values)
        for col_idx, (field_name, _) in enumerate(layout, start=1):
            cell = ws.cell(row=ridx, column=col_idx)
            cell.font = BODY_FONT
            if field_name == "state_id":
                cell.fill = fill
                cell.font = MUTED_FONT
                cell.alignment = LEFT
            elif _is_numeric_field(field_name):
                cell.alignment = RIGHT
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0"
            else:
                cell.alignment = LEFT
    _apply_widths(ws, layout)
    ws.freeze_panes = "B2"


def _write_changed_sheet(
    wb: Workbook,
    sheet_name: str,
    rows,
    layout: list[tuple[str, str]],
    ui: UI,
) -> None:
    """For the changed sheet: same column layout as the main report, but cells
    show DELTAS in place. Unchanged cells stay blank — the user can scan a row
    and immediately see which fields shifted."""
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    headers = [h for _, h in layout]
    ws.append(headers)
    _style_data_header(ws, len(headers))

    delta_col_indices: list[int] = []
    for ridx, r in enumerate(rows, start=2):
        deltas = r.deltas
        text_ctx = r.text_fields
        for col_idx, (field_name, _header) in enumerate(layout, start=1):
            cell = ws.cell(row=ridx, column=col_idx)
            cell.font = BODY_FONT
            # Identifier columns: always populate so the user knows which row
            if field_name in REGION_KEY_FIELDS:
                cell.value = r.key[REGION_KEY_FIELDS.index(field_name)] \
                    if field_name in REGION_KEY_FIELDS else None
                cell.font = MUTED_FONT
                cell.alignment = LEFT
                continue
            if field_name == "state":
                cell.value = text_ctx.get("state", "")
                cell.alignment = LEFT
                continue
            # Other columns: only fill if the field is in deltas
            if field_name not in deltas:
                continue
            ov, nv = deltas[field_name]
            if _is_numeric_field(field_name) and isinstance(ov, (int, float)) and isinstance(nv, (int, float)):
                cell.value = round(float(nv) - float(ov), 4)
                cell.number_format = "#,##0;[Red]-#,##0"
                cell.alignment = RIGHT
                if col_idx not in delta_col_indices:
                    delta_col_indices.append(col_idx)
            else:
                # Text or non-numeric change: show new value (blank if unchanged is handled by skip above)
                cell.value = nv if nv is not None else ""
                cell.alignment = LEFT

    # Conditional fills on numeric delta columns (green if positive, red if negative)
    n_rows = len(rows)
    if n_rows > 0:
        green = CellIsRule(operator="greaterThan", formula=["0"],
                           fill=PatternFill("solid", fgColor=DELTA_POSITIVE))
        red = CellIsRule(operator="lessThan", formula=["0"],
                         fill=PatternFill("solid", fgColor=DELTA_NEGATIVE))
        for col_idx in delta_col_indices:
            letter = get_column_letter(col_idx)
            rng = f"{letter}2:{letter}{n_rows + 1}"
            ws.conditional_formatting.add(rng, green)
            ws.conditional_formatting.add(rng, red)

    _apply_widths(ws, layout)
    ws.freeze_panes = "B2"


def _write_empty_sheet(
    wb: Workbook,
    sheet_name: str,
    layout: list[tuple[str, str]],
) -> None:
    """Placeholder sheet when the section had zero entries."""
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    headers = [h for _, h in layout]
    ws.append(headers)
    _style_data_header(ws, len(headers))
    ws.append(["— (none) —"] + [None] * (len(headers) - 1))
    cell = ws.cell(row=2, column=1)
    cell.font = MUTED_FONT
    cell.alignment = LEFT
    _apply_widths(ws, layout)
    ws.freeze_panes = "B2"


def write_regions_diff_xlsx(
    diff: RegionsDiffResult, path: Path, *, ui: UI | None = None,
) -> None:
    if ui is None:
        ui = get_ui("simp_chinese")
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    _write_info_sheet(wb, diff, ui)

    layout = _resolve_layout(diff)

    sections = [
        (ui["rsheet_diff_added"],   diff.added,   "full",    ADDED_FILL),
        (ui["rsheet_diff_removed"], diff.removed, "full",    REMOVED_FILL),
        (ui["rsheet_diff_changed"], diff.changed, "changed", None),
    ]
    for sheet_name, rows, kind, fill in sections:
        if not rows:
            _write_empty_sheet(wb, sheet_name, layout)
        elif kind == "full":
            _write_full_rows_sheet(wb, sheet_name, rows, layout, fill_color=fill)
        else:
            _write_changed_sheet(wb, sheet_name, rows, layout, ui)

    wb.properties.title = "V3_EAT Regions Diff"
    wb.properties.creator = "V3_EAT"
    wb.save(path)
