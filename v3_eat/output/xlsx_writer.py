from __future__ import annotations
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Iterable, Sequence

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ..analysis.construction import ConstructionRow, make_construction_columns
from ..analysis.rows import Row, make_columns
from ..i18n import UI, get_ui
from ..util.strings import BUCKET_ORDER


@dataclass
class ReportMeta:
    """Metadata describing the provenance of a report run."""
    game_version: str = ""
    raw_version: str = ""
    tool_version: str = ""
    generated_at: str = ""
    data_lang: str = ""           # game-data language used (simp_chinese / english / ...)
    ui_lang: str = ""             # UI language ("zh" or "en")
    counts: dict[str, int] = field(default_factory=dict)

    @property
    def workbook_title(self) -> str:
        v = self.raw_version or self.game_version or "unknown"
        return f"V3_EAT Report — V3 {v}"


# --- Modern palette (Tailwind-style neutrals + indigo accent) -----------------
HEADER_BG = "1F2937"
HEADER_FG = "FFFFFF"
BUILDING_BG = "EEF2FF"
BAND_BG = "F9FAFB"
BORDER_GRAY = "E5E7EB"
TEXT_PRIMARY = "111827"
TEXT_MUTED = "6B7280"

CONDITIONAL_LOW = "FCA5A5"
CONDITIONAL_MID = "FDE68A"
CONDITIONAL_HIGH = "86EFAC"

FONT_FAMILY = "Microsoft YaHei"

HEADER_FONT = Font(name=FONT_FAMILY, size=11, bold=True, color=HEADER_FG)
BODY_FONT = Font(name=FONT_FAMILY, size=10, color=TEXT_PRIMARY)
ID_FONT = Font(name=FONT_FAMILY, size=9, color=TEXT_MUTED)
BUILDING_FONT = Font(name=FONT_FAMILY, size=10, bold=True, color=TEXT_PRIMARY)
TITLE_FONT = Font(name=FONT_FAMILY, size=13, bold=True, color=TEXT_PRIMARY)

HEADER_FILL = PatternFill("solid", fgColor=HEADER_BG)
BUILDING_FILL = PatternFill("solid", fgColor=BUILDING_BG)
BAND_FILL = PatternFill("solid", fgColor=BAND_BG)

THIN = Side(style="thin", color=BORDER_GRAY)
HEADER_BORDER = Border(top=THIN, bottom=THIN)

CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", indent=1)
RIGHT = Alignment(horizontal="right", vertical="center", indent=1)

NUM_INT_KEYS = {"construction", "employment", "construction_per_lvl"}
NUM_VAL_KEYS = {
    "output_value", "input_value", "net_value",
    "material_cost_per_lvl", "wage_cost_per_lvl",
    "material_cost_per_unit", "wage_cost_per_unit", "total_cost_per_unit",
}
NUM_RATE_KEYS = {"wage_mult", "roi", "per_capita"}
TEXT_KEYS = {
    "building", "base_pms", "secondary_pms", "automation_pms", "ownership_pms",
    "building_group", "inputs_str", "outputs_str", "notes", "pm",
}
ID_KEYS = {"building_id", "base_ids", "secondary_ids", "automation_ids",
           "ownership_ids", "pm_id"}

CONDITIONAL_KEYS = ("net_value", "roi", "per_capita")
CONSTRUCTION_CONDITIONAL_KEYS = (
    "material_cost_per_unit",
    "wage_cost_per_unit",
    "total_cost_per_unit",
)

COL_WIDTH_CAP = {
    "building": 18,
    "base_pms": 24,
    "secondary_pms": 28,
    "automation_pms": 24,
    "ownership_pms": 22,
    "inputs_str": 38,
    "outputs_str": 28,
    "notes": 56,
    "building_group": 16,
    "pm": 22,
    "net_value": 18,
    "output_value": 18,
    "input_value": 18,
    "construction": 14,
    "employment": 14,
    "wage_mult": 13,
    "roi": 16,
    "per_capita": 22,
    "construction_per_lvl": 18,
    "material_cost_per_lvl": 20,
    "wage_cost_per_lvl": 28,
    "material_cost_per_unit": 22,
    "wage_cost_per_unit": 22,
    "total_cost_per_unit": 22,
}

HEADER_ONLY_WIDTH_KEYS = {
    "building_id", "base_ids", "secondary_ids", "automation_ids",
    "ownership_ids", "pm_id",
}

MIN_WIDTH = {
    "net_value": 14,
    "output_value": 14,
    "input_value": 14,
    "per_capita": 16,
    "roi": 13,
    "wage_mult": 11,
    "construction": 11,
    "employment": 11,
    "material_cost_per_lvl": 14,
    "wage_cost_per_lvl": 22,
    "material_cost_per_unit": 18,
    "wage_cost_per_unit": 16,
    "total_cost_per_unit": 18,
    "construction_per_lvl": 13,
}


def _est_width(s: str) -> int:
    n = 0
    for ch in s:
        n += 2 if ord(ch) > 127 else 1
    return n


def _cell_value(v):
    if v is None:
        return None
    if isinstance(v, float):
        if v != v:
            return None
        return round(v, 4)
    return v


def _style_data_sheet(
    ws: Worksheet,
    rows: Sequence,
    headers: list[str],
    keys: list[str],
    *,
    building_col_key: str | None = "building",
    conditional_keys: Sequence[str] = (),
    inverted_conditional_keys: Sequence[str] = (),
    fill_anchor_col: bool = True,
    data_row_offset: int = 0,
) -> None:
    """Apply modern styling to a sheet whose data rows live at
    rows [2 + data_row_offset .. 2 + data_row_offset + n_rows - 1].

    `data_row_offset` accounts for non-data rows the caller inserted between
    the header (row 1) and the actual data rows — e.g., the regions writer
    prepends a totals row at row 2, so it passes `data_row_offset=1` and
    data rows begin at row 3.
    """
    n_rows = len(rows)
    first_data_row = 2 + data_row_offset
    last_data_row = first_data_row + n_rows - 1

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = HEADER_BORDER
    ws.row_dimensions[1].height = 28

    for row_idx in range(first_data_row, last_data_row + 1):
        is_band = (row_idx % 2 == 0)
        ws.row_dimensions[row_idx].height = 18
        for col_idx, key in enumerate(keys, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if key == building_col_key:
                if fill_anchor_col:
                    cell.fill = BUILDING_FILL
                cell.font = BUILDING_FONT
                cell.alignment = CENTER
            elif key in ID_KEYS:
                if is_band:
                    cell.fill = BAND_FILL
                cell.font = ID_FONT
                cell.alignment = LEFT
            elif key in TEXT_KEYS:
                if is_band:
                    cell.fill = BAND_FILL
                cell.font = BODY_FONT
                cell.alignment = LEFT
            else:
                if is_band:
                    cell.fill = BAND_FILL
                cell.font = BODY_FONT
                cell.alignment = RIGHT

    for col_idx, key in enumerate(keys, start=1):
        letter = get_column_letter(col_idx)
        if key in NUM_INT_KEYS:
            fmt = "#,##0"
        elif key in NUM_VAL_KEYS:
            fmt = "#,##0.00;[Red]-#,##0.00"
        elif key in NUM_RATE_KEYS:
            fmt = "0.00"
        else:
            fmt = None

        header_w = _est_width(headers[col_idx - 1])
        if key in HEADER_ONLY_WIDTH_KEYS:
            ws.column_dimensions[letter].width = max(header_w + 2, 8)
            for row_idx in range(first_data_row, last_data_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if fmt is not None and isinstance(cell.value, (int, float)):
                    cell.number_format = fmt
            continue

        max_w = header_w
        for row_idx in range(first_data_row, last_data_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if fmt is not None and isinstance(cell.value, (int, float)):
                cell.number_format = fmt
            if cell.value is not None:
                w = _est_width(str(cell.value))
                if w > max_w:
                    max_w = w
        cap = COL_WIDTH_CAP.get(key, 14)
        buffer = 3 if key in NUM_VAL_KEYS or key in NUM_RATE_KEYS else 2
        width = min(max(max_w + buffer, 8), cap)
        width = max(width, MIN_WIDTH.get(key, 0))
        ws.column_dimensions[letter].width = width

    if n_rows > 0:
        rule_high_good = ColorScaleRule(
            start_type="min", start_color=CONDITIONAL_LOW,
            mid_type="percentile", mid_value=50, mid_color=CONDITIONAL_MID,
            end_type="max", end_color=CONDITIONAL_HIGH,
        )
        rule_low_good = ColorScaleRule(
            start_type="min", start_color=CONDITIONAL_HIGH,
            mid_type="percentile", mid_value=50, mid_color=CONDITIONAL_MID,
            end_type="max", end_color=CONDITIONAL_LOW,
        )
        for key in conditional_keys:
            if key not in keys:
                continue
            col_idx = keys.index(key) + 1
            letter = get_column_letter(col_idx)
            rng = f"{letter}{first_data_row}:{letter}{last_data_row}"
            ws.conditional_formatting.add(rng, rule_high_good)
        for key in inverted_conditional_keys:
            if key not in keys:
                continue
            col_idx = keys.index(key) + 1
            letter = get_column_letter(col_idx)
            rng = f"{letter}{first_data_row}:{letter}{last_data_row}"
            ws.conditional_formatting.add(rng, rule_low_good)

    if building_col_key in keys:
        ws.freeze_panes = f"{get_column_letter(keys.index(building_col_key) + 2)}2"
    else:
        ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{last_data_row}"
    ws.sheet_view.showGridLines = False


def _write_rows_sheet(wb: Workbook, sheet_name: str, rows: list[Row], ui: UI) -> Worksheet:
    ws = wb.create_sheet(sheet_name)
    columns = make_columns(ui)
    headers = [label for _, label in columns]
    keys = [k for k, _ in columns]
    ws.append(headers)
    for r in rows:
        ws.append([_cell_value(getattr(r, k)) for k in keys])
    _style_data_sheet(
        ws, rows, headers, keys,
        building_col_key="building",
        conditional_keys=CONDITIONAL_KEYS,
    )
    return ws


def _write_construction_sheet(wb: Workbook, rows: list[ConstructionRow], ui: UI) -> Worksheet:
    ws = wb.create_sheet(ui["sheet_construction"])
    columns = make_construction_columns(ui)
    headers = [label for _, label in columns]
    keys = [k for k, _ in columns]
    ws.append(headers)
    for r in rows:
        ws.append([_cell_value(getattr(r, k)) for k in keys])
    _style_data_sheet(
        ws, rows, headers, keys,
        building_col_key="building",
        conditional_keys=(),
        inverted_conditional_keys=CONSTRUCTION_CONDITIONAL_KEYS,
    )
    return ws


def _write_info_sheet(wb: Workbook, meta: ReportMeta, ui: UI) -> Worksheet:
    ws = wb.create_sheet(ui["sheet_info"], 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 38

    ws["A1"] = ui["meta_title"]
    ws["A1"].font = Font(name=FONT_FAMILY, size=14, bold=True, color=TEXT_PRIMARY)
    ws.merge_cells("A1:B1")

    rows: list[tuple[str, str]] = [
        (ui["meta_game_version"], meta.game_version or ui["meta_unknown"]),
        (ui["meta_raw_version"],  meta.raw_version or ui["meta_unknown"]),
        (ui["meta_generated_at"], meta.generated_at),
        (ui["meta_tool_version"], f"V3_EAT {meta.tool_version}"),
        (ui["meta_data_lang"],    meta.data_lang or ui["meta_unknown"]),
        (ui["meta_ui_lang"],      meta.ui_lang or ui["meta_unknown"]),
    ]
    count_keys = (
        ("count_goods",        "goods"),
        ("count_pops",         "pops"),
        ("count_pms",          "pms"),
        ("count_pmgs",         "pmgs"),
        ("count_buildings",    "buildings"),
        ("count_bgs",          "bgs"),
        ("count_combo_rows",   "combo_rows"),
        ("count_construction_rows", "construction_rows"),
    )
    for ui_key, count_key in count_keys:
        if count_key in meta.counts:
            rows.append((ui[ui_key], f"{meta.counts[count_key]:,}"))

    for i, (label, value) in enumerate(rows, start=3):
        kc = ws.cell(row=i, column=1, value=label)
        kc.font = Font(name=FONT_FAMILY, size=10, color=TEXT_MUTED)
        kc.alignment = LEFT
        vc = ws.cell(row=i, column=2, value=value)
        vc.font = Font(name=FONT_FAMILY, size=10, color=TEXT_PRIMARY, bold=True)
        vc.alignment = LEFT
        if i % 2 == 1:
            kc.fill = BAND_FILL
            vc.fill = BAND_FILL
    return ws


def write_xlsx(
    rows: Iterable[Row],
    construction_rows: Iterable[ConstructionRow],
    path: Path,
    *,
    meta: ReportMeta | None = None,
    ui: UI | None = None,
) -> int:
    rows_list = list(rows)
    construction_list = list(construction_rows)
    path.parent.mkdir(parents=True, exist_ok=True)

    if ui is None:
        ui = get_ui(meta.data_lang if meta else "simp_chinese")
    if meta is None:
        meta = ReportMeta(generated_at=datetime.now().isoformat(timespec="seconds"))
    meta.counts.setdefault("combo_rows", len(rows_list))
    meta.counts.setdefault("construction_rows", len(construction_list))

    wb = Workbook()
    default = wb.active
    wb.remove(default)

    _write_info_sheet(wb, meta, ui)

    by_bucket: dict[str, list[Row]] = defaultdict(list)
    for r in rows_list:
        by_bucket[r.bucket or "other"].append(r)

    if rows_list:
        _write_rows_sheet(wb, ui["sheet_overview"], rows_list, ui)

    for bucket in BUCKET_ORDER:
        bucket_rows = by_bucket.get(bucket, [])
        if not bucket_rows:
            continue
        bucket_rows.sort(key=lambda r: (r.net_value if r.net_value is not None else -1e18), reverse=True)
        _write_rows_sheet(wb, ui[f"bucket_{bucket}"], bucket_rows, ui)

    if construction_list:
        construction_list.sort(key=lambda r: r.construction_per_lvl)
        _write_construction_sheet(wb, construction_list, ui)

    wb.properties.title = meta.workbook_title
    wb.properties.creator = "V3_EAT"
    wb.properties.description = (
        f"Game: {meta.game_version} | Tool: V3_EAT {meta.tool_version} | "
        f"Generated: {meta.generated_at} | Data: {meta.data_lang} | UI: {meta.ui_lang}"
    )

    wb.save(path)
    return len(rows_list) + len(construction_list)
