from __future__ import annotations
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..analysis.construction import CONSTRUCTION_COLUMN_KEYS
from ..analysis.diff import (
    COMBO_KEY_FIELDS,
    COMBO_NUMERIC_FIELDS,
    COMBO_TEXT_FIELDS,
    CONSTRUCTION_KEY_FIELDS,
    CONSTRUCTION_NUMERIC_FIELDS,
    CONSTRUCTION_TEXT_FIELDS,
    ChangedRow,
    DiffResult,
)
from ..analysis.rows import COLUMN_KEYS
from ..i18n import UI, get_ui

# Identifier columns kept populated in the "changed" sheets even when no
# field-level diff falls in them, so each row is unambiguously locatable.
_COMBO_IDENTITY_FIELDS = {
    "building", "base_pms", "secondary_pms", "automation_pms",
    "ownership_pms", "building_id", "base_ids", "secondary_ids",
    "automation_ids", "ownership_ids",
}
_CONSTR_IDENTITY_FIELDS = {"building", "pm", "building_id", "pm_id"}

_COMBO_NUMERIC_SET = set(COMBO_NUMERIC_FIELDS)
_CONSTR_NUMERIC_SET = set(CONSTRUCTION_NUMERIC_FIELDS)


def _make_combo_layout(ui: UI) -> list[tuple[str, str]]:
    """Same column order as the main buildings report (analysis/rows.COLUMN_KEYS)."""
    return [(field, ui[label_key]) for field, label_key in COLUMN_KEYS]


def _make_construction_layout(ui: UI) -> list[tuple[str, str]]:
    return [(field, ui[label_key]) for field, label_key in CONSTRUCTION_COLUMN_KEYS]

HEADER_BG = "1F2937"
BAND_BG = "F9FAFB"
BORDER_GRAY = "E5E7EB"
TEXT_PRIMARY = "111827"
TEXT_MUTED = "6B7280"
DELTA_POSITIVE = "DCFCE7"
DELTA_NEGATIVE = "FEE2E2"
ADDED_FILL = "D1FAE5"
REMOVED_FILL = "FEE4E2"

FONT_FAMILY = "Microsoft YaHei"
HEADER_FONT = Font(name=FONT_FAMILY, size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name=FONT_FAMILY, size=10, color=TEXT_PRIMARY)
MUTED_FONT = Font(name=FONT_FAMILY, size=9, color=TEXT_MUTED)
TITLE_FONT = Font(name=FONT_FAMILY, size=14, bold=True, color=TEXT_PRIMARY)

HEADER_FILL = PatternFill("solid", fgColor=HEADER_BG)
BAND_FILL = PatternFill("solid", fgColor=BAND_BG)

LEFT = Alignment(horizontal="left", vertical="center", indent=1)
RIGHT = Alignment(horizontal="right", vertical="center", indent=1)
CENTER = Alignment(horizontal="center", vertical="center")


# Map canonical field name -> ui label key (mirrors COLUMN_KEYS).
_COMBO_LABEL_KEY = {
    "building_id": "col_building_id",
    "base_ids": "col_base_ids",
    "secondary_ids": "col_secondary_ids",
    "automation_ids": "col_automation_ids",
    "ownership_ids": "col_ownership_ids",
    "building": "col_building",
    "base_pms": "col_base_pms",
    "secondary_pms": "col_secondary_pms",
    "automation_pms": "col_automation_pms",
    "ownership_pms": "col_ownership_pms",
    "building_group": "col_building_group",
    "notes": "col_notes",
    "output_value": "col_output_value",
    "input_value": "col_input_value",
    "net_value": "col_net_value",
    "construction": "col_construction",
    "employment": "col_employment",
    "wage_mult": "col_wage_mult",
    "roi": "col_roi",
    "per_capita": "col_per_capita",
}

_CONSTR_LABEL_KEY = {
    "building_id": "ccol_building_id",
    "pm_id": "ccol_pm_id",
    "building": "ccol_building",
    "pm": "ccol_pm",
    "construction_per_lvl": "ccol_construction_per_lvl",
    "employment": "ccol_employment",
    "wage_mult": "ccol_wage_mult",
    "material_cost_per_lvl": "ccol_material_cost_per_lvl",
    "wage_cost_per_lvl": "ccol_wage_cost_per_lvl",
    "material_cost_per_unit": "ccol_material_cost_per_unit",
    "wage_cost_per_unit": "ccol_wage_cost_per_unit",
    "total_cost_per_unit": "ccol_total_cost_per_unit",
}


def _label(field: str, ui: UI, table: dict[str, str]) -> str:
    k = table.get(field)
    return ui[k] if k else field


def _style_header(ws) -> None:
    for c in ws[1]:
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER
    ws.row_dimensions[1].height = 24


def _write_info_sheet(wb: Workbook, diff: DiffResult, ui: UI) -> None:
    ws = wb.create_sheet(ui["sheet_info"], 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 38

    ws["A1"] = ui["diff_title"]
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:B1")

    rows = [
        (ui["diff_old_game"],      diff.old_meta.get(ui["meta_game_version"]) or diff.old_meta.get("游戏版本") or diff.old_meta.get("Game Version") or ui["meta_unknown"]),
        (ui["diff_old_generated"], diff.old_meta.get(ui["meta_generated_at"]) or diff.old_meta.get("生成时间") or diff.old_meta.get("Generated At") or ui["meta_unknown"]),
        (ui["diff_new_game"],      diff.new_meta.get(ui["meta_game_version"]) or diff.new_meta.get("游戏版本") or diff.new_meta.get("Game Version") or ui["meta_unknown"]),
        (ui["diff_new_generated"], diff.new_meta.get(ui["meta_generated_at"]) or diff.new_meta.get("生成时间") or diff.new_meta.get("Generated At") or ui["meta_unknown"]),
        (ui["diff_compared_at"],   datetime.now().isoformat(timespec="seconds")),
        ("", ""),
        (ui["diff_combo_added"],    str(len(diff.combo_added))),
        (ui["diff_combo_removed"],  str(len(diff.combo_removed))),
        (ui["diff_combo_changed"],  str(len(diff.combo_changed))),
        (ui["diff_constr_added"],   str(len(diff.construction_added))),
        (ui["diff_constr_removed"], str(len(diff.construction_removed))),
        (ui["diff_constr_changed"], str(len(diff.construction_changed))),
    ]
    for i, (k, v) in enumerate(rows, start=3):
        kc = ws.cell(row=i, column=1, value=k)
        vc = ws.cell(row=i, column=2, value=v)
        kc.font = MUTED_FONT
        vc.font = Font(name=FONT_FAMILY, size=10, bold=True, color=TEXT_PRIMARY)
        kc.alignment = LEFT
        vc.alignment = LEFT
        if i % 2 == 1:
            kc.fill = BAND_FILL
            vc.fill = BAND_FILL


def _apply_widths(ws, layout: list[tuple[str, str]]) -> None:
    for idx, (_, header) in enumerate(layout, start=1):
        letter = get_column_letter(idx)
        ws.column_dimensions[letter].width = max(10, min(36, len(str(header)) * 2 + 4))


def _write_full_diff_sheet(
    wb: Workbook,
    sheet_name: str,
    rows: list[dict],
    layout: list[tuple[str, str]],
    key_fields: tuple[str, ...],
    numeric_fields: tuple[str, ...],
    *,
    fill_color: str,
) -> None:
    """Added/removed sheet: full row data using main-table column layout."""
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    headers = [h for _, h in layout]
    ws.append(headers)
    _style_header(ws)

    fill = PatternFill("solid", fgColor=fill_color)
    key_set = set(key_fields)
    numeric_set = set(numeric_fields)

    for ridx, row in enumerate(rows, start=2):
        key = row.get("_key", ())
        for col_idx, (field, _h) in enumerate(layout, start=1):
            cell = ws.cell(row=ridx, column=col_idx)
            cell.font = BODY_FONT
            if field in key_set:
                try:
                    i = key_fields.index(field)
                    cell.value = key[i] if i < len(key) else None
                except ValueError:
                    cell.value = row.get(field)
                cell.fill = fill
                cell.font = MUTED_FONT
                cell.alignment = LEFT
            elif field in numeric_set:
                v = row.get(field)
                cell.value = v
                cell.alignment = RIGHT
                if isinstance(v, (int, float)):
                    cell.number_format = "#,##0.00"
            else:
                cell.value = row.get(field)
                cell.alignment = LEFT

    _apply_widths(ws, layout)
    ws.freeze_panes = "B2"


def _write_changed_diff_sheet(
    wb: Workbook,
    sheet_name: str,
    rows: list[ChangedRow],
    layout: list[tuple[str, str]],
    key_fields: tuple[str, ...],
    numeric_fields: tuple[str, ...],
    identity_fields: set[str],
) -> None:
    """Changed sheet: same layout as main report, in-place delta cells.
    Identifier columns (key + identity_fields) stay populated; other columns
    are blank unless that field's value changed, where:
      - numeric fields show `new - old` (color-coded)
      - text fields show new value
    """
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    headers = [h for _, h in layout]
    ws.append(headers)
    _style_header(ws)

    key_set = set(key_fields)
    numeric_set = set(numeric_fields)
    delta_col_indices: list[int] = []

    for ridx, r in enumerate(rows, start=2):
        for col_idx, (field, _h) in enumerate(layout, start=1):
            cell = ws.cell(row=ridx, column=col_idx)
            cell.font = BODY_FONT

            if field in key_set:
                try:
                    i = key_fields.index(field)
                    cell.value = r.key[i] if i < len(r.key) else None
                except ValueError:
                    pass
                cell.font = MUTED_FONT
                cell.alignment = LEFT
                continue

            if field in identity_fields:
                cell.value = r.text_fields.get(field, "")
                cell.alignment = LEFT
                continue

            if field not in r.deltas:
                continue
            ov, nv = r.deltas[field]
            if field in numeric_set and isinstance(ov, (int, float)) and isinstance(nv, (int, float)):
                cell.value = round(float(nv) - float(ov), 4)
                cell.number_format = "#,##0.00;[Red]-#,##0.00"
                cell.alignment = RIGHT
                if col_idx not in delta_col_indices:
                    delta_col_indices.append(col_idx)
            else:
                cell.value = nv if nv is not None else ""
                cell.alignment = LEFT

    n_rows = len(rows)
    if n_rows > 0 and delta_col_indices:
        green = CellIsRule(operator="greaterThan", formula=["0"],
                           fill=PatternFill("solid", fgColor=DELTA_POSITIVE))
        red = CellIsRule(operator="lessThan", formula=["0"],
                         fill=PatternFill("solid", fgColor=DELTA_NEGATIVE))
        for col_idx in delta_col_indices:
            letter = get_column_letter(col_idx)
            rng = f"{letter}2:{letter}{n_rows + 1}"
            ws.conditional_formatting.add(rng, green)
            ws.conditional_formatting.add(rng, red)

    _apply_widths(ws, layout)
    ws.freeze_panes = "B2"


def _write_empty_diff_sheet(
    wb: Workbook, sheet_name: str, layout: list[tuple[str, str]],
) -> None:
    """Placeholder sheet when a section has zero entries — explicit '— (none) —'
    notice so users don't wonder whether the diff silently failed."""
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    headers = [h for _, h in layout]
    ws.append(headers)
    _style_header(ws)
    ws.append(["— (none) —"] + [None] * (len(headers) - 1))
    cell = ws.cell(row=2, column=1)
    cell.font = MUTED_FONT
    cell.alignment = LEFT
    _apply_widths(ws, layout)
    ws.freeze_panes = "B2"


def write_diff_xlsx(diff: DiffResult, path: Path, *, ui: UI | None = None) -> None:
    if ui is None:
        ui = get_ui("simp_chinese")
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    _write_info_sheet(wb, diff, ui)

    combo_layout = _make_combo_layout(ui)
    constr_layout = _make_construction_layout(ui)

    # Sections: (sheet_ui_key, rows, layout, key_fields, numeric_fields,
    #            identity_fields, kind, fill_color)
    sections = [
        ("sheet_diff_added_combo",   diff.combo_added,   combo_layout,
         COMBO_KEY_FIELDS, COMBO_NUMERIC_FIELDS, _COMBO_IDENTITY_FIELDS, "full",    ADDED_FILL),
        ("sheet_diff_removed_combo", diff.combo_removed, combo_layout,
         COMBO_KEY_FIELDS, COMBO_NUMERIC_FIELDS, _COMBO_IDENTITY_FIELDS, "full",    REMOVED_FILL),
        ("sheet_diff_changed_combo", diff.combo_changed, combo_layout,
         COMBO_KEY_FIELDS, COMBO_NUMERIC_FIELDS, _COMBO_IDENTITY_FIELDS, "changed", None),
        ("sheet_diff_added_constr",   diff.construction_added,   constr_layout,
         CONSTRUCTION_KEY_FIELDS, CONSTRUCTION_NUMERIC_FIELDS, _CONSTR_IDENTITY_FIELDS, "full",    ADDED_FILL),
        ("sheet_diff_removed_constr", diff.construction_removed, constr_layout,
         CONSTRUCTION_KEY_FIELDS, CONSTRUCTION_NUMERIC_FIELDS, _CONSTR_IDENTITY_FIELDS, "full",    REMOVED_FILL),
        ("sheet_diff_changed_constr", diff.construction_changed, constr_layout,
         CONSTRUCTION_KEY_FIELDS, CONSTRUCTION_NUMERIC_FIELDS, _CONSTR_IDENTITY_FIELDS, "changed", None),
    ]
    for ui_key, rows, layout, kf, nf, idf, kind, fill in sections:
        sheet_name = ui[ui_key]
        if not rows:
            _write_empty_diff_sheet(wb, sheet_name, layout)
        elif kind == "full":
            _write_full_diff_sheet(wb, sheet_name, rows, layout, kf, nf, fill_color=fill)
        else:
            _write_changed_diff_sheet(wb, sheet_name, rows, layout, kf, nf, idf)

    wb.properties.title = ui["diff_title"]
    wb.properties.creator = "V3_EAT"
    wb.save(path)
