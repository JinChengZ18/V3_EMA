"""Cross-version diff: read two V3_EAT xlsx reports, compute structural and
numeric deltas, regardless of which UI language each was generated in.

The trick: each xlsx column has a translated header (e.g. "建筑" or
"Building"). We use `i18n.label_to_key()` to map any of the supported labels
back to a canonical key. Internal data is keyed by canonical names so the
diff cleanly compares mixed-language reports."""
from __future__ import annotations
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

from ..i18n import label_to_key
from ..util.logging import get_logger

log = get_logger()

# Canonical keys we care about for combo rows.
COMBO_KEY_FIELDS = ("building_id", "base_ids", "secondary_ids", "automation_ids")
COMBO_NUMERIC_FIELDS = (
    "output_value", "input_value", "net_value",
    "construction", "employment", "wage_mult", "roi", "per_capita",
)
COMBO_TEXT_FIELDS = (
    "building", "base_pms", "secondary_pms", "automation_pms",
    "ownership_pms", "building_group", "notes",
)

# Construction sheet
CONSTRUCTION_KEY_FIELDS = ("building_id", "pm_id")
CONSTRUCTION_NUMERIC_FIELDS = (
    "construction_per_lvl", "employment", "wage_mult",
    "material_cost_per_lvl", "wage_cost_per_lvl",
    "material_cost_per_unit", "wage_cost_per_unit", "total_cost_per_unit",
)
CONSTRUCTION_TEXT_FIELDS = ("building", "pm")


ComboKey = tuple[str, str, str, str]
ConstructionKey = tuple[str, str]


@dataclass
class ReportSnapshot:
    meta: dict[str, str] = field(default_factory=dict)
    combos: dict[ComboKey, dict[str, object]] = field(default_factory=dict)
    construction: dict[ConstructionKey, dict[str, object]] = field(default_factory=dict)


@dataclass
class ChangedRow:
    key: tuple
    text_fields: dict[str, str]
    deltas: dict[str, tuple[object, object]]   # field -> (old, new) for changed values only


@dataclass
class DiffResult:
    old_meta: dict[str, str]
    new_meta: dict[str, str]
    combo_added: list[dict[str, object]] = field(default_factory=list)
    combo_removed: list[dict[str, object]] = field(default_factory=list)
    combo_changed: list[ChangedRow] = field(default_factory=list)
    construction_added: list[dict[str, object]] = field(default_factory=list)
    construction_removed: list[dict[str, object]] = field(default_factory=list)
    construction_changed: list[ChangedRow] = field(default_factory=list)


def _read_info_sheet(wb) -> dict[str, str]:
    """Find the info sheet (first one is conventionally it; otherwise look by
    canonical key matched against label_to_key)."""
    candidate = None
    for name in wb.sheetnames:
        k = label_to_key(name)
        if k == "sheet_info":
            candidate = name
            break
    if candidate is None and wb.sheetnames:
        candidate = wb.sheetnames[0]
    if candidate is None:
        return {}
    ws = wb[candidate]
    out: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
        k, v = (row + (None, None))[:2]
        if k is None or v is None:
            continue
        out[str(k)] = str(v)
    return out


def _find_sheet(wb, canonical_key: str) -> str | None:
    for name in wb.sheetnames:
        if label_to_key(name) == canonical_key:
            return name
    return None


# All sheet-name canonical keys that hold combo rows (overview + all buckets).
# Reading and merging these protects against manually-edited bucket sheets:
# if a row appears in both 总览 and 基础设施 with different values, the last
# one read wins, so diff still catches the discrepancy across two reports.
_COMBO_SHEET_KEYS = (
    "sheet_overview",
    "bucket_agriculture",
    "bucket_plantations",
    "bucket_extraction",
    "bucket_manufacturing",
    "bucket_service",
    "bucket_infrastructure",
    "bucket_government",
    "bucket_military",
    "bucket_monuments",
    "bucket_other",
)


def _find_all_combo_sheets(wb) -> list[str]:
    """Return all sheet names that hold combo-row data (overview + bucket sheets),
    in priority order: overview first, then buckets. Later-read entries win."""
    canonical_set = set(_COMBO_SHEET_KEYS)
    found: list[tuple[int, str]] = []
    for name in wb.sheetnames:
        k = label_to_key(name)
        if k in canonical_set:
            # priority: overview = 0, others = 1 (so overview is read first,
            # and bucket-sheet edits OVERRIDE overview if they disagree)
            found.append((0 if k == "sheet_overview" else 1, name))
    found.sort()
    return [name for _, name in found]


def _read_data_sheet(
    wb, sheet_name: str | None,
    key_fields: tuple[str, ...],
    keep_fields: tuple[str, ...],
) -> dict[tuple, dict[str, object]]:
    """Read sheet rows into {key_tuple: {field_canonical_name: value}}.

    Header labels may be in any supported language; we use label_to_key to
    map each header back to a canonical column key (matching the keys used
    by analysis/rows.py). Columns whose label doesn't map are ignored."""
    if not sheet_name or sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]
    headers = [c.value for c in ws[1]]
    # Map header -> canonical column key by reversing the i18n table.
    # Header text -> ui_label_key like "col_building" -> field name "building"
    ui_label_keys = [label_to_key(h) if h else None for h in headers]
    # Drop "col_" / "ccol_" prefix to get the dataclass field name.
    canonical_per_col: list[str | None] = []
    for ui_k in ui_label_keys:
        if ui_k is None:
            canonical_per_col.append(None)
            continue
        # ui_k might be "col_building", "ccol_building", or "col_inputs_str", etc.
        # The convention: strip the leading "col_" or "ccol_" prefix.
        if ui_k.startswith("col_"):
            canonical_per_col.append(ui_k[4:])
        elif ui_k.startswith("ccol_"):
            canonical_per_col.append(ui_k[5:])
        else:
            canonical_per_col.append(ui_k)

    # Map each desired key/keep field to its column index.
    col_for_field: dict[str, int] = {}
    for idx, field_name in enumerate(canonical_per_col):
        if field_name and field_name not in col_for_field:
            col_for_field[field_name] = idx

    # Verify all key_fields are present.
    missing = [k for k in key_fields if k not in col_for_field]
    if missing:
        log.warning("Sheet %s missing key columns: %s (headers: %s)",
                    sheet_name, missing, headers)
        return {}

    key_indices = [col_for_field[k] for k in key_fields]
    keep_indices = {f: col_for_field[f] for f in keep_fields if f in col_for_field}

    out: dict[tuple, dict[str, object]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        key = tuple(str(row[i]) if row[i] is not None else "" for i in key_indices)
        out[key] = {f: row[idx] for f, idx in keep_indices.items()}
    return out


def read_report(path: Path) -> ReportSnapshot:
    wb = load_workbook(path, read_only=True, data_only=True)
    snap = ReportSnapshot(meta=_read_info_sheet(wb))
    # Read combo data from EVERY combo-bearing sheet (overview + buckets) and
    # merge with bucket-overrides-overview semantics, so manual edits to any
    # one sheet are still surfaced in the diff.
    combos: dict[ComboKey, dict[str, object]] = {}
    for sheet_name in _find_all_combo_sheets(wb):
        sheet_data = _read_data_sheet(
            wb, sheet_name,
            COMBO_KEY_FIELDS,
            COMBO_TEXT_FIELDS + COMBO_NUMERIC_FIELDS,
        )
        combos.update(sheet_data)   # later entries override earlier
    snap.combos = combos

    constr = _find_sheet(wb, "sheet_construction")
    snap.construction = _read_data_sheet(
        wb, constr,
        CONSTRUCTION_KEY_FIELDS,
        CONSTRUCTION_TEXT_FIELDS + CONSTRUCTION_NUMERIC_FIELDS,
    )
    wb.close()
    return snap


def _is_changed(o, n, eps_abs: float, eps_rel: float) -> bool:
    if o is None and n is None:
        return False
    if (o is None) != (n is None):
        return True
    if isinstance(o, (int, float)) and isinstance(n, (int, float)):
        diff = abs(float(o) - float(n))
        threshold = max(eps_abs, eps_rel * max(abs(float(o)), abs(float(n))))
        return diff > threshold
    return o != n


def _diff_section(
    old: dict[tuple, dict[str, object]],
    new: dict[tuple, dict[str, object]],
    text_fields: tuple[str, ...],
    numeric_fields: tuple[str, ...],
    *,
    eps_abs: float, eps_rel: float,
) -> tuple[list[dict], list[dict], list[ChangedRow]]:
    added: list[dict] = []
    removed: list[dict] = []
    changed: list[ChangedRow] = []
    old_keys = set(old)
    new_keys = set(new)
    for k in sorted(new_keys - old_keys):
        added.append({"_key": k, **new[k]})
    for k in sorted(old_keys - new_keys):
        removed.append({"_key": k, **old[k]})
    for k in sorted(old_keys & new_keys):
        o, n = old[k], new[k]
        deltas: dict[str, tuple] = {}
        for f in numeric_fields:
            ov, nv = o.get(f), n.get(f)
            if _is_changed(ov, nv, eps_abs, eps_rel):
                deltas[f] = (ov, nv)
        if deltas:
            text = {f: str(n.get(f, "")) for f in text_fields}
            changed.append(ChangedRow(key=k, text_fields=text, deltas=deltas))
    return added, removed, changed


def diff_snapshots(
    old: ReportSnapshot,
    new: ReportSnapshot,
    *,
    eps_abs: float = 0.01,
    eps_rel: float = 0.005,
) -> DiffResult:
    res = DiffResult(old_meta=old.meta, new_meta=new.meta)
    res.combo_added, res.combo_removed, res.combo_changed = _diff_section(
        old.combos, new.combos,
        COMBO_TEXT_FIELDS, COMBO_NUMERIC_FIELDS,
        eps_abs=eps_abs, eps_rel=eps_rel,
    )
    res.construction_added, res.construction_removed, res.construction_changed = _diff_section(
        old.construction, new.construction,
        CONSTRUCTION_TEXT_FIELDS, CONSTRUCTION_NUMERIC_FIELDS,
        eps_abs=eps_abs, eps_rel=eps_rel,
    )
    return res
