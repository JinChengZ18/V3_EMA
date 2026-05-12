"""Tests for V3_EMA diff workflow.

Synthetic fixtures: build two minimal V3_EMA-style xlsx reports programmatically
using the real `xlsx_writer`, then run `diff_snapshots` and assert added /
removed / changed buckets are populated correctly.

Runnable two ways:
    pytest tests/test_diff.py
    python tests/test_diff.py        (no pytest needed; falls back to plain assertions)
"""
from __future__ import annotations
import sys
import tempfile
from pathlib import Path

# Make the package importable when running as a plain script.
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from v3_ema.analysis.construction import ConstructionRow
from v3_ema.analysis.diff import diff_snapshots, read_report
from v3_ema.analysis.rows import Row
from v3_ema.output.diff_writer import write_diff_xlsx
from v3_ema.output.xlsx_writer import ReportMeta, write_xlsx


def _row(building_id: str, base_id: str, secondary_id: str, automation_id: str,
         **kwargs) -> Row:
    defaults = dict(
        building=building_id,
        base_pms=base_id, secondary_pms=secondary_id, automation_pms=automation_id,
        building_id=building_id, base_ids=base_id, secondary_ids=secondary_id,
        automation_ids=automation_id,
        bucket="测试桶",
        building_group="测试组",
    )
    defaults.update(kwargs)
    return Row(**defaults)


def _construction(building_id: str, pm_id: str, **kwargs) -> ConstructionRow:
    defaults = dict(
        building=building_id, pm=pm_id,
        building_id=building_id, pm_id=pm_id,
    )
    defaults.update(kwargs)
    return ConstructionRow(**defaults)


def _meta(version: str, data_lang: str = "simp_chinese") -> ReportMeta:
    return ReportMeta(
        game_version=version,
        raw_version=version.split(" ")[0],
        tool_version="test",
        generated_at="2026-01-01T00:00:00",
        data_lang=data_lang,
        ui_lang="zh" if data_lang == "simp_chinese" else "en",
        counts={"combo_rows": 0, "construction_rows": 0},
    )


def _meta_value(meta: dict, *labels: str) -> str | None:
    """Look up a meta value by any of the supported label spellings."""
    for lbl in labels:
        if lbl in meta:
            return meta[lbl]
    return None


def test_diff_minimal(tmp_path: Path | None = None) -> None:
    if tmp_path is None:
        tmp_path = Path(tempfile.mkdtemp())

    old_rows = [
        _row("b_a", "pm_x", "pm_n", "pm_t",
             output_value=100, input_value=20, net_value=80,
             construction=200, employment=1000,
             wage_mult=1.2, roi=0.4, per_capita=4.16),
        _row("b_b", "pm_y", "pm_n", "pm_t",
             output_value=300, input_value=50, net_value=250,
             construction=400, employment=2000,
             wage_mult=1.5, roi=0.625, per_capita=6.5),
        _row("b_dropped", "pm_z", "pm_n", "pm_t",
             output_value=10, input_value=0, net_value=10,
             construction=100, employment=500,
             wage_mult=1.0, roi=0.1, per_capita=1.04),
    ]
    old_construction = [
        _construction("b_cs", "pm_wood",
                      construction_per_lvl=2, employment=1000, wage_mult=1.5,
                      material_cost_per_lvl=1000, wage_cost_per_lvl=1500,
                      material_cost_per_unit=500, wage_cost_per_unit=750,
                      total_cost_per_unit=1250),
    ]

    new_rows = [
        _row("b_a", "pm_x", "pm_n", "pm_t",      # unchanged
             output_value=100, input_value=20, net_value=80,
             construction=200, employment=1000,
             wage_mult=1.2, roi=0.4, per_capita=4.16),
        _row("b_b", "pm_y", "pm_n", "pm_t",      # CHANGED net_value 250 -> 350
             output_value=400, input_value=50, net_value=350,
             construction=400, employment=2000,
             wage_mult=1.5, roi=0.875, per_capita=9.1),
        _row("b_added", "pm_q", "pm_n", "pm_t",  # ADDED
             output_value=500, input_value=100, net_value=400,
             construction=600, employment=3000,
             wage_mult=2.0, roi=0.667, per_capita=6.93),
        # b_dropped is REMOVED
    ]
    new_construction = [
        _construction("b_cs", "pm_wood",          # CHANGED material 500 -> 480
                      construction_per_lvl=2, employment=1000, wage_mult=1.5,
                      material_cost_per_lvl=960, wage_cost_per_lvl=1500,
                      material_cost_per_unit=480, wage_cost_per_unit=750,
                      total_cost_per_unit=1230),
    ]

    old_path = tmp_path / "old.xlsx"
    new_path = tmp_path / "new.xlsx"
    write_xlsx(old_rows, old_construction, old_path, meta=_meta("1.0.0 (Test)"))
    write_xlsx(new_rows, new_construction, new_path, meta=_meta("1.1.0 (Test)"))

    old_snap = read_report(old_path)
    new_snap = read_report(new_path)

    # Sanity: snapshots round-trip.
    assert ("b_a", "pm_x", "pm_n", "pm_t") in old_snap.combos
    assert _meta_value(old_snap.meta, "游戏版本", "Game Version") == "1.0.0 (Test)"
    assert _meta_value(new_snap.meta, "游戏版本", "Game Version") == "1.1.0 (Test)"

    diff = diff_snapshots(old_snap, new_snap)

    # --- combo assertions ---
    added_keys = {tuple(r["_key"]) for r in diff.combo_added}
    removed_keys = {tuple(r["_key"]) for r in diff.combo_removed}
    changed_keys = {tuple(r.key) for r in diff.combo_changed}

    assert added_keys == {("b_added", "pm_q", "pm_n", "pm_t")}, added_keys
    assert removed_keys == {("b_dropped", "pm_z", "pm_n", "pm_t")}, removed_keys
    assert changed_keys == {("b_b", "pm_y", "pm_n", "pm_t")}, changed_keys

    # The changed row should record the net_value delta exactly.
    [changed_row] = diff.combo_changed
    assert "net_value" in changed_row.deltas
    o, n = changed_row.deltas["net_value"]
    assert (o, n) == (250, 350), (o, n)

    # --- construction assertions ---
    [csc] = diff.construction_changed
    assert "material_cost_per_unit" in csc.deltas
    co, cn = csc.deltas["material_cost_per_unit"]
    assert (co, cn) == (500, 480)
    assert not diff.construction_added
    assert not diff.construction_removed

    # --- write diff workbook end-to-end ---
    diff_path = tmp_path / "diff.xlsx"
    write_diff_xlsx(diff, diff_path)   # default ui = zh
    assert diff_path.exists()
    # Round-trip readability
    from openpyxl import load_workbook
    wb = load_workbook(diff_path)
    assert "信息" in wb.sheetnames
    assert "新增-组合" in wb.sheetnames
    assert "移除-组合" in wb.sheetnames
    assert "变更-组合" in wb.sheetnames
    assert "变更-建造" in wb.sheetnames
    wb.close()


def test_diff_eps(tmp_path: Path | None = None) -> None:
    """Tiny floating-point drift below the absolute threshold must NOT register."""
    if tmp_path is None:
        tmp_path = Path(tempfile.mkdtemp())

    common_kwargs = dict(
        output_value=100.0, input_value=20.0,
        construction=200, employment=1000,
        wage_mult=1.2, roi=0.4, per_capita=4.16,
    )
    old_rows = [_row("b", "pm_x", "pm_n", "pm_t", net_value=80.000, **common_kwargs)]
    new_rows = [_row("b", "pm_x", "pm_n", "pm_t", net_value=80.005, **common_kwargs)]

    old_path = tmp_path / "old.xlsx"
    new_path = tmp_path / "new.xlsx"
    write_xlsx(old_rows, [], old_path, meta=_meta("v1"))
    write_xlsx(new_rows, [], new_path, meta=_meta("v2"))

    diff = diff_snapshots(read_report(old_path), read_report(new_path),
                          eps_abs=0.01, eps_rel=0.005)
    assert not diff.combo_added
    assert not diff.combo_removed
    assert not diff.combo_changed, "drift below eps must not trigger a diff"


def test_diff_eps_above_threshold(tmp_path: Path | None = None) -> None:
    """Drift above threshold MUST register."""
    if tmp_path is None:
        tmp_path = Path(tempfile.mkdtemp())

    common_kwargs = dict(
        output_value=100.0, input_value=20.0,
        construction=200, employment=1000,
        wage_mult=1.2, roi=0.4, per_capita=4.16,
    )
    old_rows = [_row("b", "pm_x", "pm_n", "pm_t", net_value=80.0, **common_kwargs)]
    new_rows = [_row("b", "pm_x", "pm_n", "pm_t", net_value=81.0, **common_kwargs)]

    old_path = tmp_path / "old.xlsx"
    new_path = tmp_path / "new.xlsx"
    write_xlsx(old_rows, [], old_path, meta=_meta("v1"))
    write_xlsx(new_rows, [], new_path, meta=_meta("v2"))

    diff = diff_snapshots(read_report(old_path), read_report(new_path))
    assert len(diff.combo_changed) == 1


def test_diff_cross_lang(tmp_path: Path | None = None) -> None:
    """Generate one report in zh UI and one in en UI; diff should still work
    (both labels map to canonical keys via i18n.label_to_key)."""
    if tmp_path is None:
        tmp_path = Path(tempfile.mkdtemp())

    common_kwargs = dict(
        output_value=100.0, input_value=20.0,
        construction=200, employment=1000,
        wage_mult=1.2, roi=0.4, per_capita=4.16,
    )
    zh_rows = [_row("b", "pm_x", "pm_n", "pm_t", net_value=80.0, **common_kwargs)]
    en_rows = [_row("b", "pm_x", "pm_n", "pm_t", net_value=90.0, **common_kwargs)]

    zh_path = tmp_path / "zh.xlsx"
    en_path = tmp_path / "en.xlsx"
    write_xlsx(zh_rows, [], zh_path, meta=_meta("zh", "simp_chinese"))
    write_xlsx(en_rows, [], en_path, meta=_meta("en", "english"))

    diff = diff_snapshots(read_report(zh_path), read_report(en_path))
    assert len(diff.combo_changed) == 1
    assert "net_value" in diff.combo_changed[0].deltas


def test_regions_diff(tmp_path: Path | None = None) -> None:
    """Build two minimal regions reports, mutate one, verify diff catches it."""
    if tmp_path is None:
        tmp_path = Path(tempfile.mkdtemp())

    from v3_ema.analysis.regions import RegionRow
    from v3_ema.analysis.regions_diff import diff_regions_snapshots, read_regions_report
    from v3_ema.output.regions_writer import write_regions_xlsx

    def mk(state_id: str, **kwargs) -> RegionRow:
        defaults = dict(
            state=state_id,
            state_id=state_id,
            strategic_region="Test Region",
            strat_id="region_test",
            arable_land=20,
            arable_buildings="Wheat",
            capped_total=50,
            capped_resources="Iron x50",
            discoverable="",
            known_resources="",
            total_capacity=70,
            traits="trait_a",
            trait_modifiers="trait_a: +0.1 throughput",
            traits_ids="state_trait_a",
            subsistence="Subsist Farm",
            provinces=10,
            numeric_id=1,
            bucket="other",
        )
        defaults.update(kwargs)
        return RegionRow(**defaults)

    old_rows = [
        mk("STATE_A", arable_land=20, capped_total=50, total_capacity=70, numeric_id=1),
        mk("STATE_B", arable_land=10, capped_total=30, total_capacity=40, numeric_id=2),
        mk("STATE_DROPPED", arable_land=5, capped_total=0, total_capacity=5, numeric_id=3),
    ]
    new_rows = [
        mk("STATE_A", arable_land=20, capped_total=50, total_capacity=70, numeric_id=1),  # unchanged
        mk("STATE_B", arable_land=15, capped_total=30, total_capacity=45, numeric_id=2),  # arable changed
        mk("STATE_NEW", arable_land=8, capped_total=20, total_capacity=28, numeric_id=4),  # added
    ]

    old_path = tmp_path / "regions_old.xlsx"
    new_path = tmp_path / "regions_new.xlsx"
    write_regions_xlsx(old_rows, old_path, meta=_meta("v1", "simp_chinese"))
    write_regions_xlsx(new_rows, new_path, meta=_meta("v2", "simp_chinese"))

    diff = diff_regions_snapshots(read_regions_report(old_path), read_regions_report(new_path))
    added_keys = {tuple(r["_key"]) for r in diff.added}
    removed_keys = {tuple(r["_key"]) for r in diff.removed}
    changed_keys = {tuple(r.key) for r in diff.changed}
    assert added_keys == {("STATE_NEW",)}, added_keys
    assert removed_keys == {("STATE_DROPPED",)}, removed_keys
    assert changed_keys == {("STATE_B",)}, changed_keys

    [chg] = diff.changed
    assert "arable_land" in chg.deltas
    assert chg.deltas["arable_land"] == (10, 15)


def test_regions_diff_cross_lang(tmp_path: Path | None = None) -> None:
    """Cross-language regions diff: same data with zh / en UIs should diff to 0."""
    if tmp_path is None:
        tmp_path = Path(tempfile.mkdtemp())

    from v3_ema.analysis.regions import RegionRow
    from v3_ema.analysis.regions_diff import diff_regions_snapshots, read_regions_report
    from v3_ema.output.regions_writer import write_regions_xlsx

    def mk_zh():
        return RegionRow(
            state="斯韦阿兰", state_id="STATE_X",
            strategic_region="北欧", strat_id="region_northern_europe",
            arable_land=30, arable_buildings="黑麦农场", capped_total=100,
            capped_resources="铁矿 ×60", discoverable="", known_resources="",
            total_capacity=130,
            traits="测试", trait_modifiers="", traits_ids="state_trait_x",
            subsistence="自给牧场", provinces=58, numeric_id=1, bucket="other",
        )

    def mk_en():
        r = mk_zh()
        # All localized text differs but canonical IDs identical
        r.state = "Svealand"
        r.strategic_region = "Northern Europe"
        r.arable_buildings = "Rye Farm"
        r.capped_resources = "Iron Mine x60"
        r.traits = "Test Trait"
        r.subsistence = "Subsistence Pasture"
        return r

    zh_path = tmp_path / "rzh.xlsx"
    en_path = tmp_path / "ren.xlsx"
    write_regions_xlsx([mk_zh()], zh_path, meta=_meta("zh", "simp_chinese"))
    write_regions_xlsx([mk_en()], en_path, meta=_meta("en", "english"))

    diff = diff_regions_snapshots(read_regions_report(zh_path), read_regions_report(en_path))
    assert not diff.added and not diff.removed and not diff.changed, \
        f"cross-lang regions diff should be empty: +{len(diff.added)}/-{len(diff.removed)}/Δ{len(diff.changed)}"


if __name__ == "__main__":
    import traceback
    failures = 0
    for name, fn in [
        ("test_diff_minimal", test_diff_minimal),
        ("test_diff_eps", test_diff_eps),
        ("test_diff_eps_above_threshold", test_diff_eps_above_threshold),
        ("test_diff_cross_lang", test_diff_cross_lang),
        ("test_regions_diff", test_regions_diff),
        ("test_regions_diff_cross_lang", test_regions_diff_cross_lang),
    ]:
        try:
            fn()
            print(f"PASS  {name}")
        except Exception as e:
            print(f"FAIL  {name}: {e}")
            traceback.print_exc()
            failures += 1
    if failures:
        sys.exit(1)
    print("All tests passed.")
